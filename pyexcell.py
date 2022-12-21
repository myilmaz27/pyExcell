

from openpyxl import Workbook,load_workbook
import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter.filedialog import askopenfile
import os

#get excell file with filedialog of tkinter

filename = filedialog.askopenfilename(initialdir="C:/", title="Select file",
                                     filetypes=(("excell files", "*.xl*"), ("all files", "*.*")))
                                          
print(filename) # check filename

wb = load_workbook(filename) # wb is Workbook
ws = wb.active   #ws is active sheet in workbook
ws2 = wb["sheetname"]  # you can create a new workksheet variable by sheet name

# Aktif çalışma sayfasının adını yazdırma
# print(wb.sheetnames)      # ['sheet1', 'sheet2', ... ]>

#print(ws)   # <Worksheet "ws sheetname">

#get single cell value by cell name
print(ws["A3"].value) 

#get single cell value by row and column number 
print("İsminiz: ", ws.cell(3,1).value)  
 
#Get one more cell value with different way
print("--------------------------------")  
print("With Range")
print("--------------------------------")
for row in range(2,5):
    for col in range(1,4):        
        print(" | " + str(ws.cell(row,col).value) + " | ",end="")
    print()
print("--------------------------------")  
print("With cell name")
print("--------------------------------")
for area in ws['A2':'C4']:
    for cel in area:
        print(" | " + str(cel.value) + " | ",end="")
    print()

print("--------------------------------")
print("Max row ve max column number")
print("--------------------------------")
for satir in range(2,ws.max_row+1):
    for sutun in range(1,ws.max_column+1):
        print(" | " + str(ws.cell(satir,sutun).value) + " | ",end="")
    print()



# Create an new Excell workbook...
new_wb = Workbook()

ws1 = new_wb.active  #select active worksheet (firt worksheet comes with workbook)
ws1.title = "First_Sheet"  #give name to activesheet
#create new worksheets with name
ws2 = new_wb.create_sheet("Second_Sheet")
ws3 = new_wb.create_sheet("Third_Sheet")

#set a value to any cell of any worksheet
ws1.cell(1,1).value="Give Value to cell(1,1) of ws1"
ws2.cell(1,1).value="Give Value to cell(1,1) of ws2"
ws3['A1']="Give Value to cell(1,1) of ws3"

print(new_wb.sheetnames)     # ['First_Sheet', 'Second_Sheet', 'Third_Sheet']
new_wb.save("pyExcell.xlsx")  # save workbook with a given name
os.system("pyExcell.xlsx")  # after save workbook you can open it